#высокоточные таблицы
from openpyxl import load_workbook
workbook = load_workbook('resources/file_example_XLSX_50.xlsx') #загрузка рабочей книги
sheet = workbook.active #активация
print(sheet.cell(row=3, column=2).value) #значение ячейки

#печать всего
for x in range(1,10):
    for y in range(1,10):
        print(sheet.cell(row=x, column=y).value)