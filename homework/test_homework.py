import os.path
import shutil
import zipfile
import csv
from zipfile import ZipFile
from PyPDF2 import PdfReader
from openpyxl import load_workbook

folder_name ='resources_hw'
pdf_file = 'docs-pytest-org-en-latest.pdf'
xlsx_file = 'file_example_XLSX_50.xlsx'
csv_file = 'SampleCSVFile_11kb.csv'
zip_file = 'zip_hw.zip'
add_file_pdf = os.path.join(folder_name, pdf_file)
add_file_xlsx = os.path.join(folder_name, xlsx_file)
add_file_csv = os.path.join(folder_name, csv_file)

def test_create_zip():
    with zipfile.ZipFile(zip_file, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        zf.write(add_file_pdf, arcname='pdf_add.pdf')
        zf.write(add_file_xlsx, arcname='xlsx_add.xlsx')
        zf.write(add_file_csv, arcname='csv_add.csv')
        zf.close()
    if os.path.exists('resources_hw/zip_hw.zip'):
        os.remove('./resources_hw/zip_hw.zip')
    shutil.move(zip_file, folder_name)  # перемещение архива

def test_read_pdf():
    with ZipFile(os.path.join(folder_name, zip_file)) as zf:
        with zf.open('pdf_add.pdf') as pdf_open:
            pdf_reader = PdfReader(pdf_open)
            assert len(pdf_reader.pages) == 412
            assert '5 Further' in pdf_reader.pages[2].extractText()


def test_read_xlsx():
    with ZipFile(os.path.join(folder_name, zip_file)) as zf:
        with zf.open('xlsx_add.xlsx') as xlsx_open:
            workbook = load_workbook(xlsx_open)
            sheet = workbook.active
            assert sheet.cell(row=3, column=2).value == 'Mara'

def test_read_csv():
    with ZipFile(os.path.join(folder_name, zip_file)) as zf:
        with open(zf.extract('csv_add.csv')) as csv_open:
            csv_f = csv.reader(csv_open)
            row_count = sum(1 for row in csv_f)
            assert row_count == 100
    os.remove('csv_add.csv')



